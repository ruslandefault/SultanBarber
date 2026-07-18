from __future__ import annotations

import io
from datetime import date, datetime, timedelta, timezone
from typing import Annotated

from fastapi import APIRouter, File, Query, UploadFile
from sqlalchemy import func, or_, select
from sqlalchemy.orm import selectinload

from app.api.deps import DbDep, UserDep
from app.core.config import settings
from app.core.errors import NotFoundError, ValidationAppError
from app.models.appointment import Appointment
from app.models.client import Client
from app.models.enums import AppointmentStatus
from app.schemas.admin import (
    AdminAppointmentOut,
    ClientCreate,
    ClientDetailOut,
    ClientImportResult,
    ClientOut,
    ClientStats,
    ClientUpdate,
)

router = APIRouter(prefix="/admin", tags=["admin:clients"])


@router.get("/clients", response_model=list[ClientOut])
async def list_clients(
    db: DbDep,
    user: UserDep,
    q: Annotated[str | None, Query(description="search by name/phone/username")] = None,
    filter: Annotated[
        str | None, Query(description="long_absent | frequent | birthday_soon")
    ] = None,
    limit: int = 100,
    offset: int = 0,
):
    salon_id = user.salon_id
    stmt = select(Client).where(Client.salon_id == salon_id)

    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            or_(
                Client.full_name.ilike(like),
                Client.phone.ilike(like),
                Client.username.ilike(like),
            )
        )

    now = datetime.now(timezone.utc)

    if filter == "long_absent":
        # clients whose last visit is > 60 days ago (or never)
        cutoff = now - timedelta(days=60)
        last_visit = (
            select(
                Appointment.client_id,
                func.max(Appointment.start_at).label("last"),
            )
            .where(Appointment.salon_id == salon_id)
            .group_by(Appointment.client_id)
            .subquery()
        )
        stmt = (
            stmt.outerjoin(last_visit, last_visit.c.client_id == Client.id)
            .where(or_(last_visit.c.last.is_(None), last_visit.c.last < cutoff))
        )
    elif filter == "frequent":
        # clients with >= 5 completed appointments
        counts = (
            select(Appointment.client_id, func.count().label("n"))
            .where(
                Appointment.salon_id == salon_id,
                Appointment.status == AppointmentStatus.completed,
            )
            .group_by(Appointment.client_id)
            .having(func.count() >= 5)
            .subquery()
        )
        stmt = stmt.join(counts, counts.c.client_id == Client.id)
    elif filter == "birthday_soon":
        # birthday within the next 14 days (month/day comparison)
        today = datetime.now(settings.tz).date()
        upcoming = [(today + timedelta(days=d)) for d in range(0, 15)]
        md = [(d.month, d.day) for d in upcoming]
        conds = [
            (func.extract("month", Client.birthday) == m)
            & (func.extract("day", Client.birthday) == d)
            for (m, d) in md
        ]
        stmt = stmt.where(Client.birthday.is_not(None)).where(or_(*conds))

    stmt = stmt.order_by(Client.full_name).limit(limit).offset(offset)
    rows = (await db.execute(stmt)).scalars().unique().all()
    return [ClientOut.model_validate(c) for c in rows]


async def _client_stats(db, salon_id: int, client_id: int) -> ClientStats:
    completed = (
        await db.execute(
            select(Appointment).where(
                Appointment.salon_id == salon_id,
                Appointment.client_id == client_id,
                Appointment.status == AppointmentStatus.completed,
            )
        )
    ).scalars().all()
    total_visits = len(completed)
    total_spend = sum(a.price_total for a in completed)
    avg_check = total_spend // total_visits if total_visits else 0
    last_visit = max((a.start_at for a in completed), default=None)
    return ClientStats(
        total_visits=total_visits,
        total_spend=total_spend,
        avg_check=avg_check,
        last_visit_at=last_visit,
    )


@router.get("/clients/{client_id}", response_model=ClientDetailOut)
async def client_detail(client_id: int, db: DbDep, user: UserDep):
    c = await db.get(Client, client_id)
    if c is None or c.salon_id != user.salon_id:
        raise NotFoundError("Mijoz topilmadi", code="client_not_found")
    history = (
        await db.execute(
            select(Appointment)
            .where(Appointment.client_id == client_id)
            .options(selectinload(Appointment.services))
            .order_by(Appointment.start_at.desc())
        )
    ).scalars().all()
    stats = await _client_stats(db, user.salon_id, client_id)
    # Build from scalar columns only — validating the Detail schema straight off the
    # ORM object triggers an async relationship lazy-load outside the greenlet (500).
    return ClientDetailOut(
        **ClientOut.model_validate(c).model_dump(),
        stats=stats,
        history=[AdminAppointmentOut.model_validate(a) for a in history],
    )


@router.post("/clients", response_model=ClientOut, status_code=201)
async def create_client(data: ClientCreate, db: DbDep, user: UserDep):
    c = Client(salon_id=user.salon_id, **data.model_dump())
    db.add(c)
    await db.commit()
    await db.refresh(c)
    return ClientOut.model_validate(c)


@router.put("/clients/{client_id}", response_model=ClientOut)
async def update_client(client_id: int, data: ClientUpdate, db: DbDep, user: UserDep):
    c = await db.get(Client, client_id)
    if c is None or c.salon_id != user.salon_id:
        raise NotFoundError("Mijoz topilmadi", code="client_not_found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    await db.commit()
    await db.refresh(c)
    return ClientOut.model_validate(c)


def _norm_phone(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    digits = "".join(ch for ch in s if ch.isdigit() or ch == "+")
    return digits or None


@router.post("/clients/import", response_model=ClientImportResult)
async def import_clients(
    db: DbDep,
    user: UserDep,
    file: Annotated[UploadFile, File(description=".xlsx with columns: full_name, phone, birthday")],
):
    """Import clients from an .xlsx file, upserting by phone.

    Expected header row (case-insensitive): full_name | name, phone, birthday (optional).
    """
    from openpyxl import load_workbook

    if not (file.filename or "").lower().endswith(".xlsx"):
        raise ValidationAppError("Faqat .xlsx fayl qabul qilinadi", code="bad_file")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValidationAppError("Faylni o'qib bo'lmadi", code="bad_xlsx") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ClientImportResult(created=0, updated=0, skipped=0)

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    def col(*names: str) -> int | None:
        for n in names:
            if n in header:
                return header.index(n)
        return None

    i_name = col("full_name", "name", "ism", "fio")
    i_phone = col("phone", "telefon", "tel")
    i_bday = col("birthday", "dob", "tugilgan_kun", "tug'ilgan_kun")

    if i_name is None or i_phone is None:
        raise ValidationAppError(
            "Sarlavhalar topilmadi: full_name va phone ustunlari kerak", code="bad_header"
        )

    # index existing clients by normalized phone
    existing = (
        await db.execute(select(Client).where(Client.salon_id == user.salon_id))
    ).scalars().all()
    by_phone = {_norm_phone(c.phone): c for c in existing if c.phone}

    created = updated = skipped = 0
    errors: list[str] = []

    for idx, row in enumerate(rows[1:], start=2):
        try:
            name = str(row[i_name]).strip() if i_name < len(row) and row[i_name] else None
            phone = _norm_phone(row[i_phone]) if i_phone < len(row) else None
            if not name or not phone:
                skipped += 1
                continue
            bday = None
            if i_bday is not None and i_bday < len(row) and row[i_bday]:
                val = row[i_bday]
                if isinstance(val, (datetime, date)):
                    bday = val.date() if isinstance(val, datetime) else val

            if phone in by_phone:
                c = by_phone[phone]
                c.full_name = name
                if bday:
                    c.birthday = bday
                updated += 1
            else:
                c = Client(
                    salon_id=user.salon_id, full_name=name, phone=phone, birthday=bday
                )
                db.add(c)
                by_phone[phone] = c
                created += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"qator {idx}: {exc}")

    await db.commit()
    return ClientImportResult(created=created, updated=updated, skipped=skipped, errors=errors)
